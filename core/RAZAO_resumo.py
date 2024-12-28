import pandas as pd
import re
import openpyxl
import os 

class RAZAO:
    def __init__(self, local, locals) -> None:
        self.local = local
        self.locals = locals

    def resumir(self):
        if os.path.isfile(self.local):
            nome = os.path.splitext(os.path.basename(self.local))[0]
            diretorio = os.path.dirname(self.local)
            return self.gerar_resumo(diretorio, self.locals, nome)


        elif os.path.isdir(self.local):
            lista = os.listdir(self.local)
            for arquivo in lista:
                if arquivo.endswith('.txt'):
                    arq = os.path.splitext(arquivo)[0]
                    self.gerar_resumo(self.local, self.locals, arq)
            else:
                return True

    def printar(self, conteudo):
        print(conteudo)

    def gerar_resumo(self, local, locals, nomearq):
        with open(f"{local}/{nomearq}.txt", "r") as arquivo:
            dados = arquivo.read()

        padrao_cabecalho = re.compile(r'LUIZA ASSESSORIA CONTABIL LTDA                    .*?Saldo', re.DOTALL)

        # Substituir o cabeçalho por uma string vazia
        texto_sem_cabecalho = re.sub(padrao_cabecalho, '', dados)
        padrao_cabecalho = re.compile(r'-{131}', re.DOTALL)
        texto_sem_cabecalho = re.sub(padrao_cabecalho, '', texto_sem_cabecalho)
        padrao_cabecalho = re.compile(r'-{88}', re.DOTALL)
        texto_sem_cabecalho = re.sub(padrao_cabecalho, '', texto_sem_cabecalho)
        padrao_cabecalho = re.compile(r'-{50}', re.DOTALL)
        texto_sem_cabecalho = re.sub(padrao_cabecalho, '', texto_sem_cabecalho)
        texto_sem_cabecalho = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]', '', texto_sem_cabecalho)
        texto_sem_cabecalho = texto_sem_cabecalho.split('\n')

        df = pd.DataFrame(columns=['Data', 'lanca', 'contra','nome','debito', 'credito', 'saldo'])
        anterior = 1

        for e, linha in enumerate(texto_sem_cabecalho):
            if 'Saldo do Mes:' in linha:
                texto_sem_cabecalho.pop(e)

        for e, linha in enumerate(texto_sem_cabecalho):
            if 'Saldo Atual:' in linha:
                texto_sem_cabecalho.pop(e)

        for e, linha in enumerate(texto_sem_cabecalho):
            if 'Saldo Anterior:' in linha:
                if anterior == 1:
                    anterior = linha
                texto_sem_cabecalho.pop(e)

        for e, linha in enumerate(texto_sem_cabecalho):
            if 'Saldo Geral:' in linha:
                texto_sem_cabecalho.pop(e)

        for e, linha in enumerate(texto_sem_cabecalho):
            if 'Sem Movimento' in linha:
                texto_sem_cabecalho.pop(e)

        for e, linha in enumerate(texto_sem_cabecalho):
            if 'Conta Custo' in linha:
                texto_sem_cabecalho.pop(e)

        for e, linha in enumerate(texto_sem_cabecalho):
            if '' == linha.strip:
                texto_sem_cabecalho.pop(e)

        for e, linha in enumerate(texto_sem_cabecalho):
            if '' == linha:
                texto_sem_cabecalho.pop(e)

        texto_sem_cabecalho = [linha for linha in texto_sem_cabecalho if not linha.isspace()]

        linhas_organizadas = re.sub(r'\n {44}', '', '\n'.join(texto_sem_cabecalho), flags=re.DOTALL)
        linhas_organizadas = linhas_organizadas.split('\n')



        for e, linha in enumerate(linhas_organizadas):
            if len(linha.split()) < 4:
                linhas_organizadas.pop(e)

        for linha in linhas_organizadas:
            linha_32 = linha[:33].strip().split()
            data = linha_32[0]
            lanca = linha_32[1]
            contra = linha_32[2]
            linha_49 = linha[-53:].replace('C', ' ').replace('D', ' ')
            if any(char.isalpha() for char in linha_49):
                debito =  ''
                credito = ''
                saldo = ''
                nome = linha[33:].strip()
            else:
                debito =  linha_49[:16].strip().replace('.', '').replace(',', '')
                credito = linha_49[16:33].strip().replace('.', '').replace(',', '')
                saldo = linha_49[33:53].strip().replace('.', '').replace(',', '')
                nome = linha[33:-49].strip()
            
            df = pd.concat([df,pd.DataFrame([{'Data': data, 'lanca': lanca, 'contra': contra,'nome': nome,'debito': debito, 'credito': credito, 'saldo': saldo}])], ignore_index=True)



        df = df.fillna(0)

        def filtarNF(x):
            padrao = r"nf\s?(n.o)?(.)?(:)?(-)?\s?(\d+)"
            try:
                return re.search(padrao, x.lower())[0].replace('nf', '').replace('n.o', '').replace('-', '').replace(':', '').replace('.', '').strip()
            except:
                try:
                    padrao = r"ft\s?(n.o)?(:)?(-)?\s?(\d+)"
                    return re.search(padrao, x.lower())[0].replace('ft', '').replace('n.o', '').replace('-', '').replace(':', '').replace('.', '').strip()
                except:
                    return '0000000'

        df['NF'] = df['nome'].apply(lambda x: int(filtarNF(x)))

        df = df[['Data', 'lanca', 'contra', 'NF', 'nome', 'debito', 'credito', 'saldo']]



        df = df.fillna(0)
        df['debito'] = pd.to_numeric(df['debito'])  
        df['credito'] = pd.to_numeric(df['credito'])  
        df['saldo'] = pd.to_numeric(df['saldo']) 
        df = df.fillna(0)


        df['debito'] = df['debito'].apply(lambda x: float('{:.2f}'.format(x / 100)))
        df['credito'] = df['credito'].apply(lambda x: float('{:.2f}'.format(x / 100)))
        df['saldo'] = df['saldo'].apply(lambda x: float('{:.2f}'.format(x / 100)))


        #filtra os df principal em df separados com base nos filtros
        df_tarifa = df[df['nome'].str.contains("TARIFA NO", case=False, na=False)].copy()
        df.drop(df_tarifa.index, inplace=True)
        df.reset_index(drop=True, inplace=True)

        df_recebimento = df[df['nome'].str.contains("Recebimento", case=False, na=False)].copy()
        df.drop(df_recebimento.index, inplace=True)
        df.reset_index(drop=True, inplace=True)

        df_venda = df[df['nome'].str.contains("VENDA DE", case=False, na=False)].copy()
        df.drop(df_venda.index, inplace=True)
        df.reset_index(drop=True, inplace=True)

        df_irrf = df[df['nome'].str.contains("RETENCAO IRRF", case=False, na=False)].copy()
        df.drop(df_irrf.index, inplace=True)
        df.reset_index(drop=True, inplace=True)

        df_pis = df[df['nome'].str.contains("RETENCAO PIS", case=False, na=False)].copy()
        df.drop(df_pis.index, inplace=True)
        df.reset_index(drop=True, inplace=True)

        df_cofins = df[df['nome'].str.contains("RETENCAO COFINS", case=False, na=False)].copy()
        df.drop(df_cofins.index, inplace=True)
        df.reset_index(drop=True, inplace=True)

        df_csll = df[df['nome'].str.contains("RETENCAO CSLL", case=False, na=False)].copy()
        df.drop(df_csll.index, inplace=True)
        df.reset_index(drop=True, inplace=True)

        df_iss = df[df['nome'].str.contains("RETENCAO ISS", case=False, na=False)].copy()
        df.drop(df_iss.index, inplace=True)
        df.reset_index(drop=True, inplace=True)

        df_inss = df[df['nome'].str.contains("RETENCAO INSS", case=False, na=False)].copy()
        df.drop(df_inss.index, inplace=True)
        df.reset_index(drop=True, inplace=True)

        


        debito_recebimento = round(df_recebimento['debito'].sum(), 2)
        credito_recebimento = round(df_recebimento['credito'].sum(), 2)

        debito_venda = round(df_venda['debito'].sum(), 2)
        credito_venda = round(df_venda['credito'].sum(), 2)

        debito_irrf = round(df_irrf['debito'].sum(), 2)
        credito_irrf = round(df_irrf['credito'].sum(), 2)

        debito_pis = round(df_pis['debito'].sum(), 2)
        credito_pis = round(df_pis['credito'].sum(), 2)

        debito_cofins = round(df_cofins['debito'].sum(), 2)
        credito_cofins = round(df_cofins['credito'].sum(), 2)

        debito_csll = round(df_csll['debito'].sum(), 2)
        credito_csll = round(df_csll['credito'].sum(), 2)

        debito_iss = round(df_iss['debito'].sum(), 2)
        credito_iss = round(df_iss['credito'].sum(), 2)

        debito_inss = round(df_inss['debito'].sum(), 2)
        credito_inss = round(df_inss['credito'].sum(), 2)

        debito_tarifa = round(df_tarifa['debito'].sum(), 2)
        credito_tarifa = round(df_tarifa['credito'].sum(), 2)


        debito_outros = round(df['debito'].sum(), 2)
        credito_outros = round(df['credito'].sum(), 2)

        total_retencoes_debito = debito_irrf + debito_pis + debito_cofins + debito_csll + debito_iss + debito_inss
        total_retencoes_credito = credito_irrf + credito_pis + credito_cofins + credito_csll + debito_iss + debito_inss

        wb = openpyxl.Workbook()

        # A primeira planilha criada automaticamente é a 'Sheet', renomeando para 'Resumo'
        ws_resumo = wb.active
        ws_resumo.title = "Resumo"

        # Adicionar dados e tabela na planilha "Resumo"
        ws_resumo['A1'] = nomearq
        ws_resumo['A5'] = 'VENDA NO MES'
        ws_resumo['A7'] = 'RETENCOES'
        ws_resumo['A8'] = 'PIS'
        ws_resumo['A9'] = 'COFINS'
        ws_resumo['A10'] = 'ISS'
        ws_resumo['A11'] = 'IRRF'
        ws_resumo['A12'] = 'CSLL'
        ws_resumo['A13'] = 'INSS'
        ws_resumo['A14'] = 'TOTAL'
        ws_resumo['A17'] = 'RECEBIMENTOS'
        ws_resumo['A18'] = 'TARIFA'
        ws_resumo['A21'] = 'TOTAL GERAL NO MES'
        ws_resumo['A25'] = 'OUTROS'

        #coluna debito
        ws_resumo['B4'] = 'ENTRADA'
        ws_resumo['B5'] = debito_venda
        ws_resumo['B8'] = debito_pis
        ws_resumo['B9'] = debito_cofins
        ws_resumo['B10'] = debito_iss
        ws_resumo['B11'] = debito_irrf
        ws_resumo['B12'] = debito_csll
        ws_resumo['B13'] = debito_inss
        ws_resumo['B14'] = total_retencoes_debito
        ws_resumo['B17'] = debito_recebimento
        ws_resumo['B18'] = debito_tarifa
        ws_resumo['B21'] = total_retencoes_debito + debito_recebimento + debito_venda + debito_tarifa
        ws_resumo['B25'] = debito_outros

        #coluna credito
        ws_resumo['C4'] = 'SAIDA'
        ws_resumo['C5'] = credito_venda
        ws_resumo['C8'] = credito_pis
        ws_resumo['C9'] = credito_cofins
        ws_resumo['C10'] = credito_iss
        ws_resumo['C11'] = credito_irrf
        ws_resumo['C12'] = credito_csll
        ws_resumo['C13'] = credito_inss
        ws_resumo['C14'] = total_retencoes_credito
        ws_resumo['C17'] = credito_recebimento
        ws_resumo['C18'] = credito_tarifa
        ws_resumo['C21'] = total_retencoes_credito + credito_recebimento +credito_venda + credito_tarifa
        ws_resumo['C25'] = credito_outros

        nomearq = f'{nomearq}-RESUMO'
        wb.save(f'{locals}/{nomearq}.xlsx')

        df_resumo = pd.read_excel(f'{locals}/{nomearq}.xlsx', sheet_name='Resumo', header=None)
        with pd.ExcelWriter(f'{locals}/{nomearq}.xlsx', engine='openpyxl') as writer:
            
            df_resumo.to_excel(writer, sheet_name='Resumo', header=None,index=False)
            # Salvando cada DataFrame em uma aba separada
            df_recebimento.to_excel(writer, sheet_name='Recebimento', index=False)
            df_venda.to_excel(writer, sheet_name='Venda de', index=False)
            df_irrf.to_excel(writer, sheet_name='Retencao IRRF', index=False)
            df_pis.to_excel(writer, sheet_name='Retencao PIS', index=False)
            df_cofins.to_excel(writer, sheet_name='Retencao COFINS', index=False)
            df_csll.to_excel(writer, sheet_name='Retencao CSLL', index=False)
            df_iss.to_excel(writer, sheet_name='Retencao ISS', index=False)
            df_inss.to_excel(writer, sheet_name='Retencao INSS', index=False)
            df_tarifa.to_excel(writer, sheet_name='Tarifa', index=False)
            df.to_excel(writer, sheet_name="OUTROS", index=False)

        # Carregar o arquivo Excel
        wb = openpyxl.load_workbook(f'{locals}/{nomearq}.xlsx')

        # Iterar por todas as planilhas
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Definir a largura de todas as colunas para 40
            for col in ws.columns:
                column = col[0].column_letter  # Obtém a letra da coluna
                ws.column_dimensions[column].width = 40  # Ajusta a largura para 40

        # Salvar o arquivo com as larguras definidas
        
        wb.save(f'{locals}/{nomearq}.xlsx')
        self.printar(f'Aquivo gerado: {nomearq}.xlsx')
        return True

class RAZAOui(RAZAO):
    def __init__(self, local, locals, ui):
        super().__init__(local, locals)
        self.ui = ui

    def printar(self, conteudo):
        self.ui.printRAZAO(conteudo)

if '__main__' == __name__:
    local = 'FAZER'
    locals = 'salvos'
    lista = os.listdir(local)
    lista = [os.path.splitext(arquivo)[0] for arquivo in lista]

    for nomearq in lista:
        RAZAO.gerar_resumo(local, locals, nomearq)